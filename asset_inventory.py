from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys

# Probeer vulnerability_counter te importeren
try:
    from vulnerability_counter import count_vulnerabilities_by_type
    VULN_COUNTER_AVAILABLE = True
except ImportError:
    VULN_COUNTER_AVAILABLE = False
    print("⚠ vulnerability_counter.py niet gevonden - CVE scan functie niet beschikbaar")

# Probeer network_scanner te importeren
try:
    from network_scanner import NetworkScanner
    NETWORK_SCANNER_AVAILABLE = True
except ImportError:
    NETWORK_SCANNER_AVAILABLE = False
    print("⚠ network_scanner.py niet gevonden - Netwerk scan functie niet beschikbaar")

def create_asset(asset_id, asset_type, ip_address, brand, model, firmware, location):
    asset = {
        'id': asset_id,
        'type': asset_type.upper(),
        'ip': ip_address,
        'brand': brand,
        'model': model,
        'firmware': firmware,
        'location': location,
        'status': 'active',
        'risk_level': 'unknown'
    }
    return asset

def display_assets(assets):
    """Toon alle assets in een geformatteerde tabel"""
    if not assets:
        print("\nGeen assets in inventaris\n")
        return
    
    print("\n" + "=" * 105)
    print(" " * 40 + "OT ASSET INVENTORY")
    print("=" * 105)
    print(f"\n{'ID':<12} {'Type':<8} {'Brand':<15} {'Model':<15} {'Firmware':<12} {'IP Address':<16} {'Location':<25}")
    print("-" * 105)
    
    for asset in assets:
        print(f"{asset['id']:<12} {asset['type']:<8} {asset['brand']:<15} {asset.get('model', 'N/A'):<15} "
              f"{asset.get('firmware', 'N/A'):<12} {asset['ip']:<16} {asset['location']:<25}")
    
    print("-" * 105)
    print(f"Totaal aantal Assets: {len(assets)}")
    
    # Statistieken
    types = {}
    for asset in assets:
        t = asset['type']
        types[t] = types.get(t, 0) + 1
    
    print(f"\nAsset Types:")
    for asset_type, count in types.items():
        print(f"  • {asset_type}: {count}")

def add_asset_interactive(assets):
    print("\n" + "=" * 50)
    print("ADD NEW ASSET".center(50))
    print("=" * 50 + "\n")
    
    asset_id = input("Asset ID (bijv. PLC-001): ").strip()
    if not asset_id:
        print("Asset ID mag niet leeg zijn")
        return
    
    if any(a['id'] == asset_id for a in assets):
        print(f"Asset {asset_id} bestaat al!")
        return
    
    print("\nTypes: PLC, HMI, RTU, SCADA, IIoT, DCS")
    asset_type = input("Type: ").strip()
    ip_address = input("IP Adres: ").strip()
    brand = input("Brand (bijv. Siemens): ").strip()
    model = input("Model (bijv. S7-1500): ").strip()
    firmware = input("Firmware versie (bijv. V2.8): ").strip()
    location = input("Locatie: ").strip()
    
    asset = create_asset(asset_id, asset_type, ip_address, brand, model, firmware, location)
    assets.append(asset)
    print(f"\n✓ Asset {asset_id} succesvol toegevoegd!")

def export_to_excel(assets, filename="assets.xlsx"):
    """
    Exporteer assets naar Excel bestand
    """
    print(f"\nExporteren naar Excel: {filename}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    headers = ['Asset ID', 'Type', 'Brand', 'Model', 'Firmware', 'IP Address', 'Location', 'Status', 'Risk Level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data toevoegen
    for idx, asset in enumerate(assets, 2):
        ws[f'A{idx}'] = asset['id']
        ws[f'B{idx}'] = asset['type']
        ws[f'C{idx}'] = asset['brand']
        ws[f'D{idx}'] = asset.get('model', 'N/A')
        ws[f'E{idx}'] = asset.get('firmware', 'N/A')
        ws[f'F{idx}'] = asset['ip']
        ws[f'G{idx}'] = asset['location']
        ws[f'H{idx}'] = asset.get('status', 'active')
        ws[f'I{idx}'] = asset.get('risk_level', 'unknown')
    
    # Kolombreedte aanpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✓ Excel bestand opgeslagen: {filename}")
    print(f"  Locatie: {os.path.abspath(filename)}")

def import_from_excel(filename="assets.xlsx"):
    """
    Importeer assets vanuit Excel bestand
    """
    if not os.path.exists(filename):
        print(f"\n❌ Bestand {filename} niet gevonden!")
        print(f"   Maak eerst een Excel bestand met optie 4 (Template aanmaken)")
        return None
    
    print(f"\nImporteren van assets uit: {filename}...")
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        assets = []
        
        # Skip header row (row 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[0]:  # Als er een Asset ID is
                try:
                    asset = {
                        'id': str(row[0]).strip(),
                        'type': str(row[1]).upper().strip() if row[1] else 'UNKNOWN',
                        'brand': str(row[2]).strip() if row[2] else 'Unknown',
                        'model': str(row[3]).strip() if row[3] else '',
                        'firmware': str(row[4]).strip() if row[4] else '',
                        'ip': str(row[5]).strip() if row[5] else 'N/A',
                        'location': str(row[6]).strip() if row[6] else 'Unknown',
                        'status': str(row[7]).strip() if row[7] else 'active',
                        'risk_level': str(row[8]).strip() if row[8] else 'unknown'
                    }
                    assets.append(asset)
                except Exception as e:
                    print(f"   ⚠ Rij {row_num} overgeslagen (fout: {e})")
        
        print(f"✓ {len(assets)} assets succesvol geïmporteerd")
        return assets
    
    except Exception as e:
        print(f"❌ Fout bij importeren: {e}")
        return None

def create_excel_template(filename="assets_template.xlsx"):
    """
    Maak een Excel template voor het importeren van assets
    """
    print(f"\nTemplate aanmaken: {filename}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ['Asset ID', 'Type', 'Brand', 'Model', 'Firmware', 'IP Address', 'Location', 'Status', 'Risk Level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Voorbeeld data
    example_data = [
        ['PLC-001', 'PLC', 'Siemens', 'S7-1500', 'V2.8', '192.168.1.10', 'Wind Turbine 1', 'active', 'unknown'],
        ['PLC-002', 'PLC', 'Rockwell', 'ControlLogix', 'V32.011', '192.168.1.11', 'Wind Turbine 2', 'active', 'unknown'],
        ['HMI-001', 'HMI', 'Siemens', 'TP1200', 'V15.1', '192.168.1.50', 'Control Room A', 'active', 'unknown'],
        ['RTU-001', 'RTU', 'Schneider', 'SCADAPack', '7.12', '192.168.1.60', 'Substation 1', 'active', 'unknown']
    ]
    
    for row_idx, data in enumerate(example_data, 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Kolombreedte aanpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✓ Template aangemaakt: {filename}")
    print(f"  Locatie: {os.path.abspath(filename)}")
    print(f"\n📝 Instructies:")
    print(f"  1. Open {filename} in Excel")
    print(f"  2. Vul je eigen assets in (verwijder voorbeelden)")
    print(f"  3. Sla op als 'assets.xlsx'")
    print(f"  4. Gebruik optie 3 in het menu om te importeren")

def run_active_network_scan(assets):
    """Run actieve netwerk scan"""
    if not NETWORK_SCANNER_AVAILABLE:
        print("\n❌ network_scanner.py niet gevonden!")
        print("  Zorg dat network_scanner.py in dezelfde map staat.")
        print("  Installeer ook: pip install scapy python-snap7")
        return
    
    print("\n" + "=" * 60)
    print(" " * 15 + "🔍 ACTIEVE NETWERK SCAN")
    print("=" * 60)
    
    network = input("\nNetwerk range (bijv. 192.168.1.0/24): ").strip()
    if not network:
        print("❌ Netwerk range is verplicht")
        return
    
    print("\n⚠ Deze scan stuurt actief pakketten naar devices")
    confirm = input("Wil je doorgaan? (j/n): ").strip().lower()
    if confirm != 'j':
        print("Scan geannuleerd.")
        return
    
    try:
        scanner = NetworkScanner()
        devices = scanner.active_scan(network, timeout=2)
        
        if devices:
            scanner.display_results(devices)
            
            # Vraag of devices moeten worden toegevoegd aan inventory
            add = input("\nVoeg gevonden devices toe aan inventory? (j/n): ").strip().lower()
            if add == 'j':
                new_assets = scanner.export_to_assets(devices)
                
                # Check voor duplicaten
                existing_ips = {a['ip'] for a in assets}
                unique_assets = [a for a in new_assets if a['ip'] not in existing_ips]
                
                assets.extend(unique_assets)
                print(f"✓ {len(unique_assets)} nieuwe assets toegevoegd")
                
                # Sla automatisch op naar Excel
                export_to_excel(assets, "assets.xlsx")
        else:
            print("\n⚠ Geen OT devices gevonden in dit netwerk")
            
    except Exception as e:
        print(f"\n❌ Fout tijdens scan: {e}")
        import traceback
        traceback.print_exc()

def run_passive_network_scan(assets):
    """Run passieve netwerk scan"""
    if not NETWORK_SCANNER_AVAILABLE:
        print("\n❌ network_scanner.py niet gevonden!")
        print("  Zorg dat network_scanner.py in dezelfde map staat.")
        print("  Installeer ook: pip install scapy python-snap7")
        return
    
    print("\n" + "=" * 60)
    print(" " * 15 + "👂 PASSIEVE NETWERK SCAN")
    print("=" * 60)
    
    print("\n⚠ Deze scan vereist root/admin rechten!")
    print("  Linux/Mac: sudo python asset_inventory.py")
    print("  Windows: Run als Administrator")
    
    duration = input("\nScan duur in seconden (standaard 60): ").strip()
    duration = int(duration) if duration else 60
    
    confirm = input(f"\nStart {duration}s passieve scan? (j/n): ").strip().lower()
    if confirm != 'j':
        print("Scan geannuleerd.")
        return
    
    try:
        scanner = NetworkScanner()
        devices = scanner.passive_scan(duration=duration)
        
        if devices:
            scanner.display_results(devices)
            
            # Vraag of devices moeten worden toegevoegd aan inventory
            add = input("\nVoeg gedetecteerde devices toe aan inventory? (j/n): ").strip().lower()
            if add == 'j':
                new_assets = scanner.export_to_assets(devices)
                
                # Check voor duplicaten
                existing_ips = {a['ip'] for a in assets}
                unique_assets = [a for a in new_assets if a['ip'] not in existing_ips]
                
                assets.extend(unique_assets)
                print(f"✓ {len(unique_assets)} nieuwe assets toegevoegd")
                
                # Sla automatisch op naar Excel
                export_to_excel(assets, "assets.xlsx")
        else:
            print("\n⚠ Geen OT traffic gedetecteerd")
            
    except PermissionError:
        print("\n❌ Onvoldoende rechten!")
        print("  Herstart het programma als administrator/root")
    except Exception as e:
        print(f"\n❌ Fout tijdens scan: {e}")
        import traceback
        traceback.print_exc()

def scan_siemens_plc(assets):
    """Scan specifieke Siemens PLC met authenticatie"""
    if not NETWORK_SCANNER_AVAILABLE:
        print("\n❌ network_scanner.py niet gevonden!")
        return
    
    print("\n" + "=" * 60)
    print(" " * 15 + "🔐 SIEMENS PLC SCAN")
    print("=" * 60)
    
    ip = input("\nPLC IP adres: ").strip()
    if not ip:
        print("❌ IP adres is verplicht")
        return
    
    rack = input("Rack nummer (standaard 0): ").strip()
    rack = int(rack) if rack else 0
    
    slot = input("Slot nummer (standaard 1): ").strip()
    slot = int(slot) if slot else 1
    
    password = input("Password (laat leeg als geen): ").strip()
    password = password if password else None
    
    try:
        scanner = NetworkScanner()
        plc_info = scanner.scan_siemens_with_auth(ip, password=password, rack=rack, slot=slot)
        
        if plc_info:
            print("\n✓ PLC informatie succesvol opgehaald")
            
            # Maak asset van PLC
            asset_id = plc_info.get('serial', f"PLC-{ip.replace('.', '-')}")
            
            asset = {
                'id': asset_id,
                'type': 'PLC',
                'ip': ip,
                'brand': plc_info.get('vendor', 'Siemens'),
                'model': plc_info.get('model', 'Unknown'),
                'firmware': plc_info.get('firmware', 'Unknown'),
                'location': input("\nLocatie van deze PLC: ").strip() or 'Unknown',
                'status': plc_info.get('status', 'active'),
                'risk_level': 'unknown'
            }
            
            # Check voor duplicaat
            if any(a['id'] == asset_id for a in assets):
                print(f"\n⚠ Asset {asset_id} bestaat al, wordt geüpdatet")
                # Update bestaande asset
                for i, a in enumerate(assets):
                    if a['id'] == asset_id:
                        assets[i] = asset
                        break
            else:
                assets.append(asset)
                print(f"\n✓ Asset {asset_id} toegevoegd")
            
            # Auto-save
            export_to_excel(assets, "assets.xlsx")
        else:
            print("\n❌ Kon geen informatie ophalen van PLC")
            print("  Check IP adres, rack/slot nummers en netwerk connectiviteit")
            
    except Exception as e:
        print(f"\n❌ Fout tijdens PLC scan: {e}")
        import traceback
        traceback.print_exc()

def run_vulnerability_scan(assets):
    """Run vulnerability scan op alle assets"""
    if not assets:
        print("\n⚠ Geen assets om te scannen!")
        print("  Voeg eerst assets toe of importeer ze uit Excel.")
        return
    
    if not VULN_COUNTER_AVAILABLE:
        print("\n❌ vulnerability_counter.py niet gevonden!")
        print("  Zorg dat vulnerability_counter.py in dezelfde map staat.")
        print("  Installeer ook: pip install requests openpyxl")
        return
    
    print("\n" + "=" * 60)
    print(" " * 15 + "🔍 VULNERABILITY SCAN STARTEN")
    print("=" * 60)
    print(f"\nTe scannen assets: {len(assets)}")
    print("Dit kan enkele minuten duren...")
    
    confirm = input("\nWil je doorgaan? (j/n): ").strip().lower()
    if confirm != 'j':
        print("Scan geannuleerd.")
        return
    
    try:
        # Run de vulnerability scan
        vuln_counts = count_vulnerabilities_by_type(assets, export_excel=True)
        
        print("\n" + "=" * 60)
        print("✓ Scan voltooid!")
        print("  - CVE rapport gegenereerd")
        print("  - Excel bestand: vulnerability_report.xlsx")
        print("=" * 60)
        
    except Exception as e:
            print(f"\n❌ Fout tijdens scan: {e}")
            print("  Controleer of alle dependencies geïnstalleerd zijn:")
            print("  pip install requests openpyxl")
    
    def search_asset(assets):
        """Zoek een asset op ID"""
        if not assets:
            print("\nGeen assets in inventaris")
            return
        
        search_term = input("\nZoek Asset ID: ").strip().upper()
        
        found = [a for a in assets if search_term in a['id'].upper()]
        
        if found:
            print(f"\n{len(found)} asset(s) gevonden:")
            display_assets(found)
        else:
            print(f"\nGeen assets gevonden met '{search_term}'")

def main():
    # Start met lege lijst (kan later geïmporteerd worden)
    assets = []
    
    # Probeer automatisch te laden bij opstarten
    if os.path.exists("assets.xlsx"):
        print("\n📂 assets.xlsx gevonden!")
        load_choice = input("Wil je automatisch laden? (j/n): ").strip().lower()
        if load_choice == 'j':
            imported = import_from_excel("assets.xlsx")
            if imported:
                assets = imported
    
    while True:
        print("\n" + "=" * 55)
        print("       OT ASSET INVENTORY - MENU")
        print("=" * 55)
        print(" 1. Toon alle assets")
        print(" 2. Nieuwe asset toevoegen (handmatig)")
        print(" 3. Importeer assets uit Excel")
        print(" 4. Exporteer assets naar Excel")
        print(" 5. Maak Excel template")
        print(" 6. Zoek asset")
        print(" 7. 🔍 Run Vulnerability Scan (CVE)")
        print(" 8. 📡 Actieve Netwerk Scan")
        print(" 9. 👂 Passieve Netwerk Scan")
        print(" 10. 🔐 Scan Siemens PLC (met authenticatie)")
        print(" 11. Afsluiten")
        print("=" * 55)
        
        choice = input("\nSelecteer optie (1-11): ").strip()
        
        if choice == '1':
            display_assets(assets)
            
        elif choice == '2':
            add_asset_interactive(assets)
            
        elif choice == '3':
            imported = import_from_excel("assets.xlsx")
            if imported:
                # Vraag of bestaande assets moeten worden vervangen of samengevoegd
                if assets:
                    print("\nWat wil je doen?")
                    print("1. Vervang huidige assets")
                    print("2. Voeg toe aan huidige assets")
                    merge_choice = input("Keuze (1/2): ").strip()
                    
                    if merge_choice == '1':
                        assets = imported
                        print("✓ Assets vervangen")
                    elif merge_choice == '2':
                        # Voeg alleen nieuwe assets toe (geen duplicaten)
                        existing_ids = {a['id'] for a in assets}
                        new_assets = [a for a in imported if a['id'] not in existing_ids]
                        assets.extend(new_assets)
                        print(f"✓ {len(new_assets)} nieuwe assets toegevoegd")
                else:
                    assets = imported
                    
        elif choice == '4':
            if not assets:
                print("\n⚠ Geen assets om te exporteren!")
                print("  Voeg eerst assets toe of importeer ze uit Excel.")
            else:
                filename = input("\nBestandsnaam (druk Enter voor 'assets.xlsx'): ").strip()
                if not filename:
                    filename = "assets.xlsx"
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                export_to_excel(assets, filename)
                
        elif choice == '5':
            filename = input("\nBestandsnaam (druk Enter voor 'assets_template.xlsx'): ").strip()
            if not filename:
                filename = "assets_template.xlsx"
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            create_excel_template(filename)
            
        elif choice == '6':
            search_asset(assets)
            
        elif choice == '7':
            run_vulnerability_scan(assets)
            
        elif choice == '8':
            run_active_network_scan(assets)
            
        elif choice == '9':
            run_passive_network_scan(assets)
            
        elif choice == '10':
            scan_siemens_plc(assets)
            
        elif choice == '11':
            print(f"\n👋 Tot ziens! Totaal beheerd: {len(assets)} assets\n")
            break
            
        else:
            print("\n❌ Ongeldige keuze. Kies 1-11.")

# Zorg dat dit HELEMAAL links staat (geen spaties ervoor)
if __name__ == "__main__":
    main()